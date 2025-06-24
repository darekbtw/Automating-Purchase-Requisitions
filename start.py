import pandas as pd
import os
from itertools import combinations
from datetime import timedelta
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def get_last_name(name):
    """Extract last name from various name formats with debug printing"""
    if not isinstance(name, str):
        return ''
    
    if ',' in name:
        last_name = name.split(',')[0].strip().upper()
        return last_name
    last_name = name.split()[-1].strip().upper()
    return last_name

def find_sum_combinations(transactions, target_sum, max_combo_size=5):
    """Find combinations of transactions that sum to target amount, with stricter matching"""
    if abs(target_sum) <= 0.001:  # Stricter zero check
        return []
    
    sorted_trans = sorted(transactions, key=lambda x: abs(x['amount']))
    n = len(sorted_trans)
    results = []
    
    for size in range(1, min(max_combo_size + 1, n + 1)):
        for combo in combinations(range(n), size):
            combo_sum = sum(sorted_trans[i]['amount'] for i in combo)
            if abs(abs(combo_sum) - abs(target_sum)) < 0.001:  # Stricter matching tolerance
                results.append([sorted_trans[i]['index'] for i in combo])
                
        if results:
            break
            
    return results

def find_matching_groups(bank_df, certify_df):
    """Strict 1-to-1 matching, then sum/combination matching up to 3 items."""
    matched_groups = []
    processed_bank_indices = set()
    processed_certify_indices = set()

    bank_df['LAST_NAME'] = bank_df['ACC.ACCOUNT NAME'].apply(get_last_name)
    certify_df['LAST_NAME'] = certify_df['Employee'].apply(get_last_name)

    bank_df['AMOUNT'] = bank_df['FIN.TRANSACTION AMOUNT'].round(2)
    certify_df['AMOUNT'] = certify_df['USD Amt'].round(2)

    max_combo_size = 3
    tolerance = 0.01

    for last_name in bank_df['LAST_NAME'].unique():
        bank_group = bank_df[(bank_df['LAST_NAME'] == last_name) & (~bank_df.index.isin(processed_bank_indices))].copy()
        certify_group = certify_df[(certify_df['LAST_NAME'] == last_name) & (~certify_df.index.isin(processed_certify_indices))].copy()

        if bank_group.empty or certify_group.empty:
            continue

        # --- Strict 1-to-1 matching ---
        for bank_idx, bank_row in bank_group.iterrows():
            if bank_idx in processed_bank_indices:
                continue
            bank_amt = bank_row['AMOUNT']
            match = certify_group[(abs(certify_group['AMOUNT'] - bank_amt) < tolerance) & (~certify_group.index.isin(processed_certify_indices))]
            if not match.empty:
                certify_idx = match.index[0]
                matched_groups.append({
                    'bank_indices': [bank_idx],
                    'certify_indices': [certify_idx],
                    'amount': bank_amt,
                    'last_name': last_name
                })
                processed_bank_indices.add(bank_idx)
                processed_certify_indices.add(certify_idx)
                print(f"DEBUG STRICT MATCH: Bank idx {bank_idx} | {bank_row['FIN.TRANSACTION DESCRIPTION']} | Amount: {bank_amt} <---> Certify idx {certify_idx} | {certify_group.loc[certify_idx]['Vendor']} | Amount: {certify_group.loc[certify_idx]['AMOUNT']}")

        # --- Combination matching: Bank to Certify ---
        # Refresh unmatched
        bank_group = bank_df[(bank_df['LAST_NAME'] == last_name) & (~bank_df.index.isin(processed_bank_indices))].copy()
        certify_group = certify_df[(certify_df['LAST_NAME'] == last_name) & (~certify_df.index.isin(processed_certify_indices))].copy()
        for bank_idx, bank_row in bank_group.iterrows():
            if bank_idx in processed_bank_indices:
                continue
            bank_amt = bank_row['AMOUNT']
            certify_indices = list(certify_group.index)
            found = False
            for size in range(2, min(max_combo_size, len(certify_indices)) + 1):
                for combo in combinations(certify_indices, size):
                    combo_sum = certify_group.loc[list(combo), 'AMOUNT'].sum()
                    if abs(combo_sum - bank_amt) < tolerance:
                        matched_groups.append({
                            'bank_indices': [bank_idx],
                            'certify_indices': list(combo),
                            'amount': bank_amt,
                            'last_name': last_name
                        })
                        processed_bank_indices.add(bank_idx)
                        processed_certify_indices.update(combo)
                        print(f"DEBUG COMBO MATCH: Bank idx {bank_idx} | {bank_row['FIN.TRANSACTION DESCRIPTION']} | Amount: {bank_amt} <---> Certify idxs {combo} | Amounts: {list(certify_group.loc[list(combo), 'AMOUNT'])}")
                        found = True
                        break
                if found:
                    break

        # --- Combination matching: Certify to Bank ---
        # Refresh unmatched
        bank_group = bank_df[(bank_df['LAST_NAME'] == last_name) & (~bank_df.index.isin(processed_bank_indices))].copy()
        certify_group = certify_df[(certify_df['LAST_NAME'] == last_name) & (~certify_df.index.isin(processed_certify_indices))].copy()
        for certify_idx, certify_row in certify_group.iterrows():
            if certify_idx in processed_certify_indices:
                continue
            certify_amt = certify_row['AMOUNT']
            bank_indices = list(bank_group.index)
            found = False
            for size in range(2, min(max_combo_size, len(bank_indices)) + 1):
                for combo in combinations(bank_indices, size):
                    combo_sum = bank_group.loc[list(combo), 'AMOUNT'].sum()
                    if abs(combo_sum - certify_amt) < tolerance:
                        matched_groups.append({
                            'bank_indices': list(combo),
                            'certify_indices': [certify_idx],
                            'amount': certify_amt,
                            'last_name': last_name
                        })
                        processed_bank_indices.update(combo)
                        processed_certify_indices.add(certify_idx)
                        print(f"DEBUG COMBO MATCH: Certify idx {certify_idx} | {certify_row['Vendor']} | Amount: {certify_amt} <---> Bank idxs {combo} | Amounts: {list(bank_group.loc[list(combo), 'AMOUNT'])}")
                        found = True
                        break
                if found:
                    break

    return matched_groups

def remove_zero_sum_groups(df, name_col, amount_col, desc_col=None, date_col=None, max_group_size=4):
    """Remove groups of transactions that sum to zero for each person"""
    result_df = df.copy()
    removed_indices = set()
    
    def find_zero_sum_groups(transactions):
        zero_sum_groups = set()
        # First look for direct positive/negative pairs
        amounts_dict = {}
        for idx, row in transactions.iterrows():
            amount = round(row[amount_col], 2)  # Round to 2 decimal places
            if abs(amount) >= 0.01:  # Ignore very small amounts
                if amount in amounts_dict:
                    amounts_dict[amount].append(idx)
                else:
                    amounts_dict[amount] = [idx]
        # Match exact opposite amounts
        for amount in list(amounts_dict.keys()):
            if -amount in amounts_dict:
                pos_indices = amounts_dict[amount]
                neg_indices = amounts_dict[-amount]
                min_pairs = min(len(pos_indices), len(neg_indices))
                zero_sum_groups.update(pos_indices[:min_pairs])
                zero_sum_groups.update(neg_indices[:min_pairs])
        # Then try combinations for any remaining amounts
        remaining_indices = [idx for idx in transactions.index if idx not in zero_sum_groups]
        if remaining_indices:
            for size in range(2, min(max_group_size + 1, len(remaining_indices) + 1)):
                for combo in combinations(remaining_indices, size):
                    combo_sum = sum(transactions.loc[i, amount_col] for i in combo)
                    if abs(round(combo_sum, 2)) <= 0.01:  # Round sum before checking
                        zero_sum_groups.update(combo)
        return zero_sum_groups
    # Process each person's transactions
    for name in df[name_col].unique():
        person_mask = df[name_col] == name
        person_transactions = df[person_mask]
        # If date column provided, group by date first
        if date_col:
            for date in person_transactions[date_col].unique():
                date_mask = person_transactions[date_col] == date
                date_transactions = person_transactions[date_mask]
                # Find zero-sum groups within this date
                zero_sum_indices = find_zero_sum_groups(date_transactions)
                removed_indices.update(zero_sum_indices)
        else:
            # Find zero-sum groups for all person's transactions
            zero_sum_indices = find_zero_sum_groups(person_transactions)
            removed_indices.update(zero_sum_indices)
    # Remove the identified zero-sum groups
    if removed_indices:
        # Print details of removed transactions
        removed_df = result_df.loc[list(removed_indices)]
        print("\nRemoving zero-sum transactions:")
        for _, row in removed_df.iterrows():
            print(f"{row[name_col]}: {row[amount_col]} - {row[desc_col] if desc_col else ''}")
        result_df = result_df.drop(index=removed_indices)
    return result_df

def reconcile_statements(bank_file_path, certify_file_path):
    """Reconciliation that preserves all original data"""
    
    # Read original files
    bank_df = pd.read_excel(bank_file_path)
    certify_df = pd.read_excel(certify_file_path)
    
    # Remove BILLING ACCOUNT entries and RBT transactions with negative values
    bank_df = bank_df[bank_df['ACC.ACCOUNT NAME'] != 'BILLING ACCOUNT']
    
    # Create RBT mask for entries with RBT at start or end and negative values
    rbt_mask = (
        ~(
            (
                bank_df['FIN.TRANSACTION DESCRIPTION'].str.contains(r'^RBT\s|^RBT$|\sRBT\s*$', case=False, regex=True, na=False)
            ) &
            (bank_df['FIN.TRANSACTION AMOUNT'] < 0)  # Negative values only
        )
    )
    bank_df = bank_df[rbt_mask]
    
    # Remove zero amounts
    bank_df = bank_df[abs(bank_df['FIN.TRANSACTION AMOUNT']) >= 0.01]
    certify_df = certify_df[abs(certify_df['USD Amt']) >= 0.01]
    
    # Remove zero-sum groups from both datasets
    print("\nChecking for zero-sum transaction groups...")
    bank_df = remove_zero_sum_groups(
        bank_df,
        name_col='ACC.ACCOUNT NAME',
        amount_col='FIN.TRANSACTION AMOUNT',
        desc_col='FIN.TRANSACTION DESCRIPTION',
        date_col='FIN.POSTING DATE'
    )
    
    certify_df = remove_zero_sum_groups(
        certify_df,
        name_col='Employee',
        amount_col='USD Amt',
        desc_col='Vendor',
        date_col='Processed Date'
    )
    
    print("\nStarting reconciliation process...")
    print(f"Total bank transactions: {len(bank_df)}")
    print(f"Total certify transactions: {len(certify_df)}")
    
    # Find matching groups
    matched_groups = find_matching_groups(bank_df, certify_df)
    
    matches = []
    matched_bank_indices = set()
    matched_certify_indices = set()
    
    # Process each matched group, checking for zero-sum groups
    for group in matched_groups:
        # Skip if the group's transactions sum to zero
        bank_sum = bank_df.loc[group['bank_indices']]['FIN.TRANSACTION AMOUNT'].sum().round(2)
        if abs(bank_sum) <= 0.01:
            print(f"Skipping zero-sum group with bank total: {bank_sum}")
            continue
        
        bank_entries = bank_df.loc[group['bank_indices']]
        certify_entries = certify_df.loc[group['certify_indices']]
        
        # Calculate group total for reference
        group_total = bank_entries['FIN.TRANSACTION AMOUNT'].sum().round(2)
        
        # For each certify transaction in the group, output a row with all details and the group total (bank sum)
        bank_descriptions = "; ".join(bank_entries['FIN.TRANSACTION DESCRIPTION'].astype(str))
        bank_dates = "; ".join(bank_entries['FIN.POSTING DATE'].astype(str))
        for _, certify_row in certify_entries.iterrows():
            matches.append({
                'Last Name': group['last_name'],
                'Amount': certify_row['AMOUNT'],  # Always Certify amount
                'Bank Date': bank_dates,  # Concatenated bank dates
                'Bank Description': bank_descriptions,
                'Certify Description': certify_row['Vendor'],
                'Expense Category': certify_row['Expense Category'],
                'Group Total': group_total  # This is the sum of the matched bank transactions
            })
        
        # Track matched indices
        matched_bank_indices.update(group['bank_indices'])
        matched_certify_indices.update(group['certify_indices'])
    
    # Get unmatched entries using indices
    unmatched_bank = bank_df.loc[~bank_df.index.isin(matched_bank_indices)]
    unmatched_certify = certify_df.loc[~certify_df.index.isin(matched_certify_indices)]
    
    # Remove zero-sum groups from unmatched entries
    print("\nChecking for zero-sum groups in unmatched entries...")
    
    if not unmatched_bank.empty:
        original_unmatched_bank = len(unmatched_bank)
        unmatched_bank = remove_zero_sum_groups(
            unmatched_bank,
            name_col='ACC.ACCOUNT NAME',
            amount_col='FIN.TRANSACTION AMOUNT',
            desc_col='FIN.TRANSACTION DESCRIPTION',
            date_col='FIN.POSTING DATE'
        )
        removed_bank = original_unmatched_bank - len(unmatched_bank)
        if removed_bank > 0:
            print(f"Removed {removed_bank} bank transactions that formed zero-sum groups")
    
    if not unmatched_certify.empty:
        original_unmatched_certify = len(unmatched_certify)
        unmatched_certify = remove_zero_sum_groups(
            unmatched_certify,
            name_col='Employee',
            amount_col='USD Amt',
            desc_col='Vendor',
            date_col='Processed Date'
        )
        removed_certify = original_unmatched_certify - len(unmatched_certify)
        if removed_certify > 0:
            print(f"Removed {removed_certify} certify transactions that formed zero-sum groups")
    
    print(f"\nFound {len(matched_groups)} matching groups")
    print(f"Matched bank transactions: {len(matched_bank_indices)}")
    print(f"Matched certify transactions: {len(matched_certify_indices)}")
    print(f"Final unmatched bank transactions: {len(unmatched_bank)}")
    print(f"Final unmatched certify transactions: {len(unmatched_certify)}")
    
    return matches, unmatched_bank, unmatched_certify

def save_results(matches, unmatched_bank_df, unmatched_certify_df, output_dir="reconciliation_output"):
    """Save results with all original columns preserved"""
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Save matches
    matches_df = pd.DataFrame(matches)
    if not matches_df.empty:
        matches_df = matches_df.sort_values(['Last Name', 'Group Total', 'Amount'])
        # Add Mismatch column
        matches_df['Mismatch'] = matches_df.apply(lambda row: '' if abs(row['Amount'] - row['Group Total']) < 0.01 else 'Mismatch', axis=1)
    match_path = os.path.join(output_dir, "matched_transactions.xlsx")
    matches_df.to_excel(match_path, index=False)
    
    # Post-process with openpyxl for SUM row and coloring mismatches
    wb = load_workbook(match_path)
    ws = wb.active
    nrows = ws.max_row
    ncols = ws.max_column
    # Add SUM row for B and G
    sum_row = nrows + 1
    ws[f"A{sum_row}"] = "SUM:"
    ws[f"B{sum_row}"] = f"=SUM(B2:B{nrows})"
    ws[f"G{sum_row}"] = f"=SUM(G2:G{nrows})"
    # Color mismatches in H
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    for row in range(2, nrows+1):
        mismatch_cell = ws[f"H{row}"]
        if mismatch_cell.value == "Mismatch":
            mismatch_cell.fill = red_fill
    wb.save(match_path)
    
    # Save unmatched transactions with all original columns
    unmatched_bank_df.to_excel(os.path.join(output_dir, "unmatched_bank.xlsx"), index=False)
    unmatched_certify_df.to_excel(os.path.join(output_dir, "unmatched_certify.xlsx"), index=False)
    
    # Print summary
    print("\nReconciliation Summary:")
    print(f"Matched Transactions: {len(matches_df)}")
    print(f"Unmatched Bank Transactions: {len(unmatched_bank_df)}")
    print(f"Unmatched Certify Transactions: {len(unmatched_certify_df)}")
    print("\nFiles saved in:", output_dir)

class ModernReconciliationGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Certify Reconciliation Tool")
        self.root.geometry("800x600")
        
        # Configure colors
        self.BLUE = "#184287"
        self.COPPER = "#cd916d"
        self.BG_COLOR = "#ffffff"
        self.HOVER_COPPER = "#d9a68c"
        
        # File paths
        self.bank_file_path = tk.StringVar()
        self.certify_file_path = tk.StringVar()
        self.selected_year = tk.StringVar()
        self.selected_month = tk.StringVar()
        
        # Configure styles
        self.setup_styles()
        self.create_gui()
        
    def setup_styles(self):
        # Configure custom styles
        style = ttk.Style()
        
        # Set default font to Arial
        self.root.option_add("*Font", "Arial 10")
        
        # Configure frame style
        style.configure("Modern.TFrame", background=self.BG_COLOR)
        
        # Configure label styles
        style.configure("Title.TLabel", 
                       font=('Arial', 24, 'bold'), 
                       foreground=self.BLUE,
                       background=self.BG_COLOR)
        
        style.configure("Subtitle.TLabel", 
                       font=('Arial', 14, 'bold'),
                       foreground=self.BLUE,
                       background=self.BG_COLOR)
        
        style.configure("Modern.TLabel",
                       font=('Arial', 10),
                       foreground=self.BLUE,
                       background=self.BG_COLOR)
        
        # Configure button styles
        style.configure("Modern.TButton",
                       font=("Arial", 10),
                       foreground=self.BLUE,
                       background=self.COPPER,
                       padding=(10, 5))
        
        style.map("Modern.TButton",
                 foreground=[('active', self.BLUE)],
                 background=[('active', self.HOVER_COPPER)])
        
        # Configure entry style
        style.configure("Modern.TEntry",
                       fieldbackground="white",
                       foreground=self.BLUE,
                       padding=(5, 5))
        
        # Configure combobox style
        style.configure("Modern.TCombobox",
                       fieldbackground="white",
                       foreground=self.BLUE,
                       padding=(5, 5))
        
    def create_gui(self):
        # Main container
        main_frame = ttk.Frame(self.root, style="Modern.TFrame", padding="20")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        # Configure grid weight
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, 
                              text="Certify Reconciliation Tool",
                              style="Title.TLabel")
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 30), sticky="w")
        
        # File Selection Section
        file_section = ttk.Frame(main_frame, style="Modern.TFrame")
        file_section.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 20))
        
        ttk.Label(file_section, 
                 text="File Selection",
                 style="Subtitle.TLabel").grid(row=0, column=0, columnspan=3, pady=(0, 15), sticky="w")
        
        # Bank File
        ttk.Label(file_section, 
                 text="Bank Statement:",
                 style="Modern.TLabel").grid(row=1, column=0, sticky="w")
        ttk.Entry(file_section,
                 textvariable=self.bank_file_path,
                 style="Modern.TEntry",
                 width=50).grid(row=1, column=1, padx=10)
        ttk.Button(file_section,
                  text="Browse",
                  style="Modern.TButton",
                  command=lambda: self.browse_file('bank')).grid(row=1, column=2)
        
        # Certify File
        ttk.Label(file_section,
                 text="Certify Report:",
                 style="Modern.TLabel").grid(row=2, column=0, sticky="w", pady=(15, 0))
        ttk.Entry(file_section,
                 textvariable=self.certify_file_path,
                 style="Modern.TEntry",
                 width=50).grid(row=2, column=1, padx=10, pady=(15, 0))
        ttk.Button(file_section,
                  text="Browse",
                  style="Modern.TButton",
                  command=lambda: self.browse_file('certify')).grid(row=2, column=2, pady=(15, 0))
        
        # Date Selection Section
        date_section = ttk.Frame(main_frame, style="Modern.TFrame")
        date_section.grid(row=2, column=0, columnspan=3, sticky="ew", pady=20)
        
        ttk.Label(date_section,
                 text="Output File Naming",
                 style="Subtitle.TLabel").grid(row=0, column=0, columnspan=3, pady=(0, 15), sticky="w")
        
        # Year Selection
        ttk.Label(date_section,
                 text="Year:",
                 style="Modern.TLabel").grid(row=1, column=0, sticky="w")
        years = [str(year) for year in range(2020, datetime.now().year + 2)]
        year_combo = ttk.Combobox(date_section,
                                 textvariable=self.selected_year,
                                 values=years,
                                 style="Modern.TCombobox",
                                 width=15)
        year_combo.grid(row=1, column=1, sticky="w", padx=(10, 0))
        year_combo.set(str(datetime.now().year))
        
        # Month Selection
        ttk.Label(date_section,
                 text="Month:",
                 style="Modern.TLabel").grid(row=2, column=0, sticky="w", pady=(15, 0))
        months = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
        month_combo = ttk.Combobox(date_section,
                                  textvariable=self.selected_month,
                                  values=months,
                                  style="Modern.TCombobox",
                                  width=15)
        month_combo.grid(row=2, column=1, sticky="w", padx=(10, 0), pady=(15, 0))
        month_combo.set(datetime.now().strftime('%m'))
        
        # Process Button
        process_button = ttk.Button(main_frame,
                                  text="Process Reconciliation",
                                  style="Modern.TButton",
                                  command=self.process_reconciliation)
        process_button.grid(row=3, column=0, columnspan=3, pady=30)
        
        # Status Label
        self.status_label = ttk.Label(main_frame,
                                    text="",
                                    style="Modern.TLabel")
        self.status_label.grid(row=4, column=0, columnspan=3, pady=10)

    def browse_file(self, file_type):
        filename = filedialog.askopenfilename(
            title=f"Select {file_type.capitalize()} File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            if file_type == 'bank':
                self.bank_file_path.set(filename)
            else:
                self.certify_file_path.set(filename)
    
    def process_reconciliation(self):
        if not self.bank_file_path.get() or not self.certify_file_path.get():
            messagebox.showerror("Error", "Please select both bank and certify files.")
            return
            
        if not self.selected_year.get() or not self.selected_month.get():
            messagebox.showerror("Error", "Please select both year and month.")
            return
            
        try:
            self.status_label.config(text="Processing... Please wait.")
            self.root.update()
            
            # Create output directory with year and month
            output_dir = f"reconciliation_{self.selected_year.get()}_{self.selected_month.get()}"
            
            # Process reconciliation (using your existing functions)
            matches, unmatched_bank, unmatched_certify = reconcile_statements(
                self.bank_file_path.get(),
                self.certify_file_path.get()
            )
            
            # Save results
            save_results(matches, unmatched_bank, unmatched_certify, output_dir)
            
            self.status_label.config(text="✓ Reconciliation completed successfully!")
            messagebox.showinfo("Success", f"Reconciliation completed! Files saved in {output_dir}")
            
        except Exception as e:
            self.status_label.config(text="✗ Error occurred during processing!")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        # Command line mode remains unchanged
        current_dir = os.path.dirname(os.path.abspath(__file__))
        bank_file = os.path.join(current_dir, "bank_statement.xlsx")
        certify_file = os.path.join(current_dir, "certify_report.xlsx")
        
        try:
            matches, unmatched_bank, unmatched_certify = reconcile_statements(
                bank_file, certify_file
            )
            save_results(matches, unmatched_bank, unmatched_certify)
        except FileNotFoundError as e:
            print(f"\nError: File not found - {str(e)}")
            print("Please make sure both bank_statement.xlsx and certify_report.xlsx exist in the program directory.")
        except Exception as e:
            print(f"\nUnexpected error occurred: {str(e)}")
            print("If this error persists, please check the file formats and contents.")
        finally:
            print("\nProgram execution completed.")
    else:
        # GUI mode with new modern interface
        try:
            root = tk.Tk()
            app = ModernReconciliationGUI(root)
            root.mainloop()
        except Exception as e:
            print(f"Error starting GUI: {str(e)}")
            sys.exit(1)
        print("\nProgram execution completed.")